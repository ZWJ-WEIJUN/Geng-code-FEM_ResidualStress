# -*- coding: utf-8 -*-
"""
Created on Fri May 31 11:32:31 2026

@author: Weijun
"""

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
plt.close('all')
plt.rcParams["font.family"] = "Times New Roman"
plt.rcParams["font.size"] = 18
plt.rcParams["axes.titlesize"] = 24
plt.rcParams["axes.labelsize"] = 24
plt.rcParams["axes.labelweight"] = "bold"
plt.rcParams["xtick.labelsize"] = 20
plt.rcParams["ytick.labelsize"] = 20


WORKBOOK_PATH = '/Users/zhangweijun/Documents/GitHub/Geng-code-FEM_ResidualStress/CompCtrl_DataExtraction_Width/DataFile/WP_All_Width_03102026.xlsx'
SHEET_NAME = 'Sheet1'



def clean_data_v1(workbook_path, sheet_name, block_name):
      # Variables definition and initialization
    # For one wall
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    header_col = None
    for cell in ws[2]:
        if isinstance(cell.value, str) and cell.value.strip() == block_name:
            header_col = cell.column
            break

    if header_col is None:
        raise ValueError(f"Could not find block name '{block_name}' on row 2")

    x_col = None
    y_col = None
    for offset in range(0, 8):
        cell = ws.cell(row=3, column=header_col + offset)
        value = str(cell.value).strip().lower() if cell.value is not None else ''
        if value in ('x_shifted', 'x-shifted', 'x shifted'):
            x_col = header_col + offset
            y_candidate = ws.cell(row=3, column=x_col + 1)
            if y_candidate.value is not None and str(y_candidate.value).strip().lower() == 'y':
                y_col = x_col + 1
                break

    if x_col is None or y_col is None:
        raise ValueError(f"Could not find X_Shifted/Y columns for '{block_name}'")

    print(f"Workbook sheet: {sheet_name}, block: {block_name}")
    print(f"Header column: {header_col}, X column: {x_col}, Y column: {y_col}")

    cleaned_out_lines = []
    invalid_lines = []
    for row_num, row in enumerate(
        ws.iter_rows(min_row=4, min_col=x_col, max_col=y_col, values_only=True), start=4
    ):
        x_value = row[0]
        y_value = row[1]
        if x_value is None or y_value is None:
            continue
        try:
            float(x_value)
            float(y_value)
            cleaned_out_lines.append(f"{x_value},{y_value}")
        except (ValueError, TypeError):
            invalid_lines.append((row_num, x_value, y_value))

    print(f"Found {len(cleaned_out_lines)} valid numerical data lines")
    print(f"Skipped {len(invalid_lines)} non-numerical lines")

    wb.close()
    return cleaned_out_lines

def width_cal(cleanData,layer_target,y_range,y_inc):
    
    print(f"width_cal called with {len(cleanData)} data points")
    
    # *********************************************************************************************************************
    # Start of the data analysis
    # *********************************************************************************************************************
    # Create a 2D plot for STW70_LP
    df = pd.DataFrame([x.split(',') for x in cleanData]) # Convert the data into a DataFrame
    
    print(f"DataFrame created with shape: {df.shape}")
    
    df = df.map(str.strip)  # Remove leading and trailing whitespace characters
    

    df = df.astype(float)  # Convert the data into float
    # print(df)
    df.iloc[:, 1] = df.iloc[:, 1].abs()
    df_original = df.copy()
    # plt.show()
    y_min = 0.5
    y_max = 55
    
    # y_range = 15
    # y_inc = 0.5
    
    df =df[(df[1] > y_min) & (df[1] < y_max)]
    
    left_x = df[0].min()
    right_x =df[0].max()
    mid_x = (left_x+right_x)/2
    
    df_left = df[(df[0] > left_x-2) & (df[0] < left_x+2)]
    df_right = df[(df[0] > right_x-2) & (df[0] < right_x+2)]
    
    print(f"Total filtered data points: {len(df)}")
    print(f"Left side points: {len(df_left)}")
    print(f"Right side points: {len(df_right)}")
    print(f"Left + Right = {len(df_left) + len(df_right)}")
    print(f"Left X range: {left_x-2} to {left_x+2}")
    print(f"Right X range: {right_x-2} to {right_x+2}")
    
    if (len(df_left)+len(df_right)!=len(df)):
        print(f"WARNING: Data points mismatch! Total: {len(df)}, Left+Right: {len(df_left)+len(df_right)}")
        print("Some points may fall between left and right regions or be duplicated.")
        print("Continuing with available data...")
        # quit()  # Commenting out quit() to continue execution
    
    left_x_mean = []
    right_x_mean = []    
    
    
    Height_info = []
  
    for i in layer_target:
        for j in np.arange(i, i+y_range,y_inc):
            Height_info.append(j)
            
            # Get points in the current height range
            left_points = df_left[(df_left[1] > j) & (df_left[1] < j+y_inc)]
            right_points = df_right[(df_right[1] > j) & (df_right[1] < j+y_inc)]
            
            if len(left_points) == 0:
                cal_left = np.nan
                print(f"Warning: No left points found for height {j:.1f}")
            else:
                cal_left = left_points[0].mean()
            
            if len(right_points) == 0:
                cal_right = np.nan
                print(f"Warning: No right points found for height {j:.1f}")
            else:
                cal_right = right_points[0].mean()
            
            left_x_mean.append(cal_left)
            right_x_mean.append(cal_right)
    
    left_x_mean = np.array(left_x_mean)
    right_x_mean = np.array(right_x_mean)
    Width = right_x_mean - left_x_mean 
    
    # Calculate mean and std ignoring nan values
    Width_mean = np.nanmean(Width)
    valid_width = Width[~np.isnan(Width)]
    Width_SD = np.nanstd(valid_width, ddof=1) if len(valid_width) > 1 else np.nan
    
    print(f"Width calculation complete. Valid points: {np.sum(~np.isnan(Width))}/{len(Width)}")
    print(f"Width mean: {Width_mean:.4f}, SD: {Width_SD:.4f}")
    
    return Width,Height_info, Width_mean, Width_SD, df_original, mid_x


def calculate_ci(values):
    valid_values = np.asarray(values, dtype=float)
    valid_values = valid_values[~np.isnan(valid_values)]
    n_points = len(valid_values)

    if n_points == 0:
        return np.nan, np.nan, np.nan, np.nan, 0

    mean = np.mean(valid_values)
    std = np.std(valid_values, ddof=1) if n_points > 1 else np.nan
    sem = std / np.sqrt(n_points) if n_points > 1 else np.nan
    ci_margin = 1.96 * sem if n_points > 1 else np.nan

    return mean, std, mean - ci_margin, mean + ci_margin, n_points

def plot_graph(dataframes, mid_x_values, colors, markers, alphas, labels):
    fig, ax = plt.subplots(figsize=(11, 8.5), dpi=300)

    reference_mid_x = mid_x_values[0]

    for df, mid_x, color, marker, alpha, label in zip(dataframes, mid_x_values, colors, markers, alphas, labels):
        df_x = df.iloc[:, 0]
        df_y = df.iloc[:, 1]
        x_offset = mid_x - reference_mid_x
        adjusted_x = df_x - x_offset - 10
        ax.scatter(adjusted_x, df_y, color=color, marker=marker, alpha=alpha, label=label, s=5)

    ax.axvline(x=-1.575, color='blue', linestyle='--', linewidth=2.0, label='Target Wall Edge')
    ax.axvline(x=1.575, color='blue', linestyle='--', linewidth=2.0)

    ax.set_xlabel('Width - X (mm)', fontweight="bold", fontsize=24)
    ax.set_ylabel('Height - Z (mm)', fontweight="bold", fontsize=24)
    
    # Make tick labels bold
    for label in ax.get_xticklabels():
        label.set_fontweight("bold")
    for label in ax.get_yticklabels():
        label.set_fontweight("bold")
    
    ax.tick_params(axis='both', labelcolor='black', labelsize=20, length=7, width=1.5)
    plt.xlim(-2.5, 2.5)
    plt.ylim(0, 55)
    plt.legend(scatterpoints=1, loc='lower center', ncol=1, fontsize=18, markerscale=3)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig('/Users/zhangweijun/Documents/GitHub/Geng-code-FEM_ResidualStress/CompCtrl_DataExtraction_Width/Created_Figures/TW234_SideProfile_NoCtrl_Comparison.pdf', bbox_inches='tight')
    


if __name__ == '__main__':
    y_range = 55
    y_inc = 1.0
    height_target = [0.5]

    tw_configs = [
        ('TW2 - No Ctrl (0.8 mm)', 'WP10_#1_CMM_Cam_B73', 'black', 'o', 1.0),
        ('TW3 - PFR Ctrl', 'WP9_#2_CMM_Cam_B70', 'grey', 'o', 1.0),
        ('TW4 - QPF Height Ctrl', 'WP9_#3_CMM_Cam_B40', 'tab:blue', 'o', 1.0),
    ]

    results = []
    print('Data loading from Excel sheet complete. Start width calculation\n')

    for tw_label, block_name, color, marker, alpha in tw_configs:
        print(f'Starting width calculation for {tw_label}...')
        clean_data = clean_data_v1(WORKBOOK_PATH, SHEET_NAME, block_name)
        width_values, height_info, width_mean, width_sd, df_raw, mid_x = width_cal(clean_data, height_target, y_range, y_inc)
        results.append({
            'label': tw_label,
            'color': color,
            'marker': marker,
            'alpha': alpha,
            'width_values': width_values,
            'height_info': height_info,
            'width_mean': width_mean,
            'width_sd': width_sd,
            'df_raw': df_raw,
            'mid_x': mid_x,
        })
        print(f'Completed width calculation for {tw_label}')

    print('All width calculations completed!')

    for result in results:
        print(f"{result['label']} Mean: {result['width_mean']}")
        print(f"{result['label']} SD: {result['width_sd']}")

    summary_rows = []
    for result in results:
        mean, std, ci_lower, ci_upper, n_points = calculate_ci(result['width_values'])
        summary_rows.append((result['label'], mean, std, ci_lower, ci_upper, n_points))

    print()
    print("=" * 110)
    print("Summary Table")
    print("=" * 110)
    print(f"{'Thin Wall':<55s} | {'Mean (mm)':>10s} | {'Std Dev (mm)':>12s} | {'95% CI Lower':>12s} | {'95% CI Upper':>12s} | {'N Points':>8s}")
    print("-" * 110)
    for label, mean, std, ci_lower, ci_upper, n_points in summary_rows:
        print(f"{label:<55s} | {mean:>10.4f} | {std:>12.4f} | {ci_lower:>12.4f} | {ci_upper:>12.4f} | {n_points:>8d}")
    print()

    dataframes = [result['df_raw'] for result in results]
    mid_x_values = [result['mid_x'] for result in results]
    colors = [result['color'] for result in results]
    markers = [result['marker'] for result in results]
    alphas = [result['alpha'] for result in results]
    labels = [result['label'] for result in results]

    plot_graph(dataframes, mid_x_values, colors, markers, alphas, labels)

    fig2, ax2 = plt.subplots(figsize=(11, 8.5), dpi=300)
    for result in results:
        ax2.scatter(
            result['height_info'],
            result['width_values'],
            color=result['color'],
            marker=result['marker'],
            alpha=result['alpha'],
            label=result['label'],
        )
    ax2.axhline(y=3.15, color='blue', linestyle='--', linewidth=2.0, label='Target Lower Limit (3.15 mm)')
    ax2.axhline(y=3.25, color='blue', linestyle='--', linewidth=2.0, label='Target Upper Limit (3.25 mm)')
    ax2.set_ylim(2.4, 4.5)
    ax2.set_xlabel('Height (mm)', fontweight='bold', fontsize=24)
    ax2.set_ylabel('Width (mm)', fontweight='bold', fontsize=24)
    ax2.tick_params(axis='both', labelcolor='black', labelsize=20, length=7, width=1.5)
    for lbl in ax2.get_xticklabels():
        lbl.set_fontweight('bold')
    for lbl in ax2.get_yticklabels():
        lbl.set_fontweight('bold')
    ax2.legend(scatterpoints=1, loc='upper left', ncol=1, fontsize=18, markerscale=1.5)
    ax2.grid(True, alpha=0.3)
    plt.tight_layout()
    fig2.savefig('/Users/zhangweijun/Documents/GitHub/Geng-code-FEM_ResidualStress/CompCtrl_DataExtraction_Width/Created_Figures/TW234_Width_Scatter.pdf', bbox_inches='tight')

    fig3, ax3 = plt.subplots(figsize=(11, 8.5), dpi=300)
    for result in results:
        for i in range(len(height_target)):
            start = i * y_range
            end = (i + 1) * y_range
            ax3.plot(
                result['height_info'][start:end],
                result['width_values'][start:end],
                color=result['color'],
                marker=result['marker'],
                alpha=result['alpha'],
                label=result['label'] if i == 0 else '',
                linewidth=3,
                markersize=8,
                linestyle='-',
            )
    ax3.axhline(y=3.15, color='blue', linestyle='--', linewidth=2.0, label='Target Lower Limit (3.15 mm)')
    ax3.axhline(y=3.25, color='blue', linestyle='--', linewidth=2.0, label='Target Upper Limit (3.25 mm)')
    ax3.set_ylim(2.4, 4.5)
    ax3.set_xlabel('Height (mm)', fontweight='bold', fontsize=24)
    ax3.set_ylabel('Width (mm)', fontweight='bold', fontsize=24)
    ax3.tick_params(axis='both', labelcolor='black', labelsize=20, length=7, width=1.5)
    for lbl in ax3.get_xticklabels():
        lbl.set_fontweight('bold')
    for lbl in ax3.get_yticklabels():
        lbl.set_fontweight('bold')
    ax3.legend(scatterpoints=1, loc='upper left', ncol=1, fontsize=18, markerscale=1.5)
    ax3.grid(True, alpha=0.3)
    plt.tight_layout()
    fig3.savefig('/Users/zhangweijun/Documents/GitHub/Geng-code-FEM_ResidualStress/CompCtrl_DataExtraction_Width/Created_Figures/TW234_Width_Line.pdf', bbox_inches='tight')

    plt.show()
